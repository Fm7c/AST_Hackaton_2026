#!/usr/bin/env python3
"""Analisa registos do sensor AS3935 e cria um Excel com o nome da pasta.

Uso:
    python analisar_relampagos.py "/caminho/para/a/pasta"
    python analisar_relampagos.py "/caminho/raiz" --cada-pasta

O modo --cada-pasta cria um Excel em cada pasta (incluindo subpastas)
que contenha ficheiros .txt diretamente no seu interior.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit(
        "Falta o pacote 'openpyxl'. Instale-o no Terminal com: "
        "python3 -m pip install openpyxl"
    ) from exc


PASTA_PREDEFINIDA = Path.home() / "Desktop" / "LightningData" / "Data" / "D05"


TIMESTAMP_RE = re.compile(
    r"^(?P<ts>\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(?:\.\d{1,6})?)$"
)
ELAPSED_RE = re.compile(r"^\[(?P<seconds>\d+)\s*s\]\s*(?P<message>.*)$", re.I)
DISTANCE_RE = re.compile(r"(?P<distance>\d+(?:[.,]\d+)?)\s*km", re.I)
TUNING_RE = re.compile(
    r"TUN_CAP\s*=\s*(?P<cap>\d+)\s+IRQ\s*=\s*(?P<irq>\d+)\s*Hz\s+"
    r"RES\s*=\s*(?P<res>\d+)\s*Hz\s+err\s*=\s*(?P<err>-?\d+)",
    re.I,
)
SETTING_PATTERNS = {
    "noise_floor": re.compile(r"noise\s*floor\s*[:=]\s*(-?\d+)", re.I),
    "spike_rejection": re.compile(r"spike\s*rejection\s*[:=]\s*(-?\d+)", re.I),
    "watchdog_threshold": re.compile(r"watchdog\s*threshold\s*[:=]\s*(-?\d+)", re.I),
    "sensibilidade_generica": re.compile(r"(?:sensibilidade|sensitivity)\s*[:=]\s*([^,;]+)", re.I),
    "modo_afe": re.compile(r"(?:afe\s*(?:mode|gain)|modo\s*afe|indoor/outdoor)\s*[:=]\s*([^,;]+)", re.I),
    "melhor_condensador": re.compile(r"best\s*capacitor\s*=\s*(-?\d+)", re.I),
    "ressonancia_hz": re.compile(r"resonance\s*=\s*(-?\d+)", re.I),
}

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

STATE_KEYS = [
    "noise_floor",
    "spike_rejection",
    "watchdog_threshold",
    "sensibilidade_generica",
    "modo_afe",
    "melhor_condensador",
    "ressonancia_hz",
]

SENSITIVITY_KEYS = [
    "noise_floor",
    "spike_rejection",
    "watchdog_threshold",
    "sensibilidade_generica",
    "modo_afe",
]


def sem_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def normalizar(texto: str) -> str:
    return re.sub(r"\s+", " ", sem_acentos(texto).strip().lower())


def parse_datetime(texto: str) -> datetime | None:
    texto = texto.strip().replace("T", " ")
    for formato in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            pass
    return None


def data_referencia_do_nome(nome: str) -> datetime | None:
    m = re.search(r"(\d{8})[_-](\d{4})h?", nome)
    if m:
        try:
            return datetime.strptime("".join(m.groups()), "%Y%m%d%H%M")
        except ValueError:
            pass

    m = re.search(
        r"(?:Thunder[_-])?(\d{1,2})(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(\d{4})[_-](\d{4})",
        nome,
        re.I,
    )
    if m:
        dia, mes_txt, ano, hora = m.groups()
        mes = MONTHS[mes_txt.lower()[:3]]
        hh, mm = int(hora[:2]), int(hora[2:])
        try:
            return datetime(int(ano), mes, int(dia), hh, mm)
        except ValueError:
            pass
    return None


def ler_linhas(caminho: Path) -> list[str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return caminho.read_text(encoding=encoding).splitlines()
        except UnicodeDecodeError:
            continue
    return caminho.read_text(errors="replace").splitlines()


def estado_vazio() -> dict[str, Any]:
    return {chave: None for chave in STATE_KEYS}


def estado_texto(estado: dict[str, Any]) -> str:
    partes: list[str] = []
    etiquetas = {
        "noise_floor": "NF",
        "spike_rejection": "SR",
        "watchdog_threshold": "WD",
        "sensibilidade_generica": "Sens",
        "modo_afe": "AFE",
        "melhor_condensador": "CAP",
        "ressonancia_hz": "RES",
    }
    for chave in SENSITIVITY_KEYS:
        valor = estado.get(chave)
        if valor not in (None, ""):
            sufixo = " Hz" if chave == "ressonancia_hz" else ""
            partes.append(f"{etiquetas[chave]}={valor}{sufixo}")
    return " | ".join(partes) if partes else "Desconhecida"


def snapshot_estado(estado: dict[str, Any]) -> dict[str, Any]:
    return {chave: estado.get(chave) for chave in STATE_KEYS}


def classificar_evento(mensagem: str) -> tuple[str | None, str | None, float | str | None]:
    n = normalizar(mensagem)

    if "lightning strike detected" in n or "relampago" in n or re.search(r"\braio\b", n):
        distancia: float | None = None
        if "menos de 1km" in n or "menos de 1 km" in n:
            distancia = "<1"
        else:
            m = DISTANCE_RE.search(mensagem)
            if m:
                distancia = float(m.group("distance").replace(",", "."))
        return "Relâmpago", mensagem.strip(), distancia

    if n in {"disturber.", "disturber"} or "disturbio detectado" in n:
        return "Distúrbio", mensagem.strip(), None

    if n in {"noise.", "noise"} or "muito ruido" in n:
        return "Ruído", mensagem.strip(), None

    return None, None, None


def e_inicio_sessao(mensagem: str) -> bool:
    n = normalizar(mensagem)
    return "franklin lightning detector" in n or "starting lc auto-tuning" in n


def e_mensagem_arranque(mensagem: str) -> bool:
    n = normalizar(mensagem)
    return e_inicio_sessao(mensagem) or "lightning detector ready" in n or "detector ready" in n


def lista_ficheiros_txt(pasta: Path) -> list[Path]:
    ficheiros = [p for p in pasta.iterdir() if p.is_file() and p.suffix.lower() == ".txt"]
    return sorted(
        ficheiros,
        key=lambda p: (data_referencia_do_nome(p.name) is None, data_referencia_do_nome(p.name) or datetime.max, p.name.lower()),
    )


def analisar_pasta(pasta: Path) -> dict[str, Any]:
    ficheiros = lista_ficheiros_txt(pasta)
    if not ficheiros:
        raise ValueError(f"A pasta não contém ficheiros .txt: {pasta}")

    eventos: list[dict[str, Any]] = []
    configuracoes: list[dict[str, Any]] = []
    desconhecidas: list[dict[str, Any]] = []
    resumo_ficheiros: list[dict[str, Any]] = []

    estado = estado_vazio()
    sessao = 0
    ordem_evento = 0
    ordem_config = 0

    for caminho in ficheiros:
        linhas = ler_linhas(caminho)
        referencia = data_referencia_do_nome(caminho.name)
        timestamp_atual: datetime | None = None
        estado_inicio = estado_texto(estado)
        contagens = Counter()
        arranques = 0
        desconhecidas_ficheiro = 0
        tem_timestamp_explicito = False
        ultimo_relampago_antigo: dict[str, Any] | None = None

        for numero_linha, linha_original in enumerate(linhas, start=1):
            linha = linha_original.strip()
            if not linha:
                continue

            m_ts = TIMESTAMP_RE.match(linha)
            if m_ts:
                timestamp_atual = parse_datetime(m_ts.group("ts"))
                tem_timestamp_explicito = True
                continue

            segundos: int | None = None
            mensagem = linha
            m_elapsed = ELAPSED_RE.match(mensagem)
            if m_elapsed:
                segundos = int(m_elapsed.group("seconds"))
                mensagem = m_elapsed.group("message").strip()

            # Distância da mensagem antiga em duas linhas.
            n = normalizar(mensagem)
            if n.startswith("approximately:") and ultimo_relampago_antigo is not None:
                m_dist = DISTANCE_RE.search(mensagem)
                if m_dist:
                    ultimo_relampago_antigo["distancia_km"] = float(m_dist.group("distance").replace(",", "."))
                ultimo_relampago_antigo["detalhe"] = (
                    f"{ultimo_relampago_antigo['detalhe']} {mensagem}".strip()
                )
                ultimo_relampago_antigo = None
                continue

            if e_mensagem_arranque(mensagem):
                if e_inicio_sessao(mensagem):
                    sessao += 1
                    estado = estado_vazio()
                    arranques += 1
                elif sessao == 0:
                    sessao = 1

                ordem_config += 1
                configuracoes.append({
                    "ordem": ordem_config,
                    "ficheiro": caminho.name,
                    "linha": numero_linha,
                    "data_hora": timestamp_atual,
                    "data_referencia": referencia,
                    "sessao": sessao,
                    "categoria": "Arranque",
                    "parametro": "Sensor",
                    "valor_anterior": None,
                    "valor_novo": mensagem,
                    "unidade": None,
                    "mensagem": mensagem,
                    "tun_cap": None,
                    "irq_hz": None,
                    "res_teste_hz": None,
                    "erro_hz": None,
                    **snapshot_estado(estado),
                    "estado_sensibilidade": estado_texto(estado),
                })
                continue

            if sessao == 0:
                sessao = 1  # sessão implícita quando o cabeçalho não está no conjunto de ficheiros

            m_tuning = TUNING_RE.search(mensagem)
            if m_tuning:
                ordem_config += 1
                configuracoes.append({
                    "ordem": ordem_config,
                    "ficheiro": caminho.name,
                    "linha": numero_linha,
                    "data_hora": timestamp_atual,
                    "data_referencia": referencia,
                    "sessao": sessao,
                    "categoria": "Auto-tuning",
                    "parametro": "Teste TUN_CAP",
                    "valor_anterior": None,
                    "valor_novo": int(m_tuning.group("cap")),
                    "unidade": None,
                    "mensagem": mensagem,
                    "tun_cap": int(m_tuning.group("cap")),
                    "irq_hz": int(m_tuning.group("irq")),
                    "res_teste_hz": int(m_tuning.group("res")),
                    "erro_hz": int(m_tuning.group("err")),
                    **snapshot_estado(estado),
                    "estado_sensibilidade": estado_texto(estado),
                })
                continue

            alteracao_encontrada = False
            for chave, padrao in SETTING_PATTERNS.items():
                m_setting = padrao.search(mensagem)
                if not m_setting:
                    continue
                alteracao_encontrada = True
                anterior = estado.get(chave)
                novo_txt = m_setting.group(1).strip()
                novo: Any = int(novo_txt) if re.fullmatch(r"-?\d+", novo_txt) else novo_txt
                estado[chave] = novo
                ordem_config += 1
                configuracoes.append({
                    "ordem": ordem_config,
                    "ficheiro": caminho.name,
                    "linha": numero_linha,
                    "data_hora": timestamp_atual,
                    "data_referencia": referencia,
                    "sessao": sessao,
                    "categoria": "Sensibilidade" if chave in {
                        "noise_floor", "spike_rejection", "watchdog_threshold",
                        "sensibilidade_generica", "modo_afe"
                    } else "Calibração",
                    "parametro": chave,
                    "valor_anterior": anterior,
                    "valor_novo": novo,
                    "unidade": "Hz" if chave == "ressonancia_hz" else None,
                    "mensagem": mensagem,
                    "tun_cap": None,
                    "irq_hz": None,
                    "res_teste_hz": None,
                    "erro_hz": None,
                    **snapshot_estado(estado),
                    "estado_sensibilidade": estado_texto(estado),
                })
                break
            if alteracao_encontrada:
                continue

            tipo, detalhe, distancia = classificar_evento(mensagem)
            if tipo:
                ordem_evento += 1
                evento = {
                    "ordem": ordem_evento,
                    "ficheiro": caminho.name,
                    "linha": numero_linha,
                    "data_hora": timestamp_atual,
                    "data_referencia": referencia,
                    "segundos": segundos,
                    "tipo": tipo,
                    "detalhe": detalhe,
                    "distancia_km": distancia,
                    "sessao": sessao,
                    **snapshot_estado(estado),
                    "estado_sensibilidade": estado_texto(estado),
                }
                eventos.append(evento)
                contagens[tipo] += 1
                ultimo_relampago_antigo = evento if normalizar(mensagem) == "lightning strike detected!" else None
                continue

            # A linha que apenas anuncia o início de auto-tuning pode ser registada sem ser desconhecida.
            if "auto-tuning" in n:
                ordem_config += 1
                configuracoes.append({
                    "ordem": ordem_config,
                    "ficheiro": caminho.name,
                    "linha": numero_linha,
                    "data_hora": timestamp_atual,
                    "data_referencia": referencia,
                    "sessao": sessao,
                    "categoria": "Auto-tuning",
                    "parametro": "Início",
                    "valor_anterior": None,
                    "valor_novo": None,
                    "unidade": None,
                    "mensagem": mensagem,
                    "tun_cap": None,
                    "irq_hz": None,
                    "res_teste_hz": None,
                    "erro_hz": None,
                    **snapshot_estado(estado),
                    "estado_sensibilidade": estado_texto(estado),
                })
                continue

            desconhecidas_ficheiro += 1
            desconhecidas.append({
                "ficheiro": caminho.name,
                "linha": numero_linha,
                "data_hora": timestamp_atual,
                "data_referencia": referencia,
                "sessao": sessao,
                "conteudo": mensagem,
            })

        resumo_ficheiros.append({
            "ficheiro": caminho.name,
            "linhas": len(linhas),
            "timestamp_explicito": "Sim" if tem_timestamp_explicito else "Não",
            "data_referencia": referencia,
            "arranques": arranques,
            "relampagos": contagens["Relâmpago"],
            "disturbios": contagens["Distúrbio"],
            "ruido": contagens["Ruído"],
            "total_deteccoes": sum(contagens.values()),
            "sensibilidade_inicio": estado_inicio,
            "sensibilidade_fim": estado_texto(estado),
            "linhas_nao_reconhecidas": desconhecidas_ficheiro,
        })

    return {
        "pasta": pasta,
        "ficheiros": ficheiros,
        "eventos": eventos,
        "configuracoes": configuracoes,
        "desconhecidas": desconhecidas,
        "resumo_ficheiros": resumo_ficheiros,
        "sessoes": max([e["sessao"] for e in eventos] + [c["sessao"] for c in configuracoes] + [0]),
    }



# -----------------------------------------------------------------------------
# Criação do Excel com openpyxl (compatível com macOS/Windows/Linux)
# -----------------------------------------------------------------------------

AZUL_ESCURO = "17365D"
AZUL_CABECALHO = "1F4E78"
BRANCO = "FFFFFF"
BORDA = "B8C4CE"


def valor_excel(valor: Any) -> Any:
    """Converte valores que o Excel/openpyxl não aceita diretamente."""
    if isinstance(valor, Path):
        return str(valor)
    return valor


def aplicar_cabecalho(ws: Any, linha: int, col_inicio: int, col_fim: int) -> None:
    fill = PatternFill("solid", fgColor=AZUL_CABECALHO)
    font = Font(bold=True, color=BRANCO)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style="thin", color=BORDA)
    border = Border(left=side, right=side, top=side, bottom=side)
    for col in range(col_inicio, col_fim + 1):
        cell = ws.cell(linha, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[linha].height = 28


def aplicar_bordas(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    side = Side(style="thin", color="D6DEE5")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def definir_larguras(ws: Any, larguras: dict[str, float]) -> None:
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura


def criar_tabela(ws: Any, referencia: str, nome: str) -> None:
    tabela = Table(displayName=nome, ref=referencia)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)


def escrever_linhas(ws: Any, cabecalhos: list[str], linhas: list[list[Any]]) -> int:
    ws.append(cabecalhos)
    aplicar_cabecalho(ws, 1, 1, len(cabecalhos))
    for linha in linhas:
        ws.append([valor_excel(v) for v in linha])
    return ws.max_row


def criar_excel(resultado: dict[str, Any], destino: Path) -> Path:
    eventos = resultado["eventos"]
    configs = resultado["configuracoes"]
    ficheiros = resultado["resumo_ficheiros"]
    desconhecidas = resultado["desconhecidas"]

    wb = Workbook()
    resumo = wb.active
    resumo.title = "Resumo"
    sh_eventos = wb.create_sheet("Relampagos")
    sh_cont = wb.create_sheet("Contagens")
    sh_sens = wb.create_sheet("Sensibilidade")
    sh_fich = wb.create_sheet("Ficheiros")
    sh_desc = wb.create_sheet("Nao_reconhecidas") if desconhecidas else None

    # ---------------- Resumo ----------------
    resumo.merge_cells("A1:F1")
    titulo = resumo["A1"]
    titulo.value = f"Análise de relâmpagos — {resultado['pasta'].name}"
    titulo.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    titulo.font = Font(bold=True, color=BRANCO, size=16)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    resumo.row_dimensions[1].height = 34

    resumo.append([])
    resumo.append(["Indicador", "Valor"])
    aplicar_cabecalho(resumo, 3, 1, 2)
    indicadores = [
        ["Pasta analisada", str(resultado["pasta"])],
        ["Ficheiros .txt", len(ficheiros)],
        ["Sessões identificadas", resultado["sessoes"]],
        ["Relâmpagos", sum(f["relampagos"] for f in ficheiros)],
        ["Distúrbios", sum(f["disturbios"] for f in ficheiros)],
        ["Ruído", sum(f["ruido"] for f in ficheiros)],
        ["Total de deteções", sum(f["total_deteccoes"] for f in ficheiros)],
        ["Linhas não reconhecidas", len(desconhecidas)],
    ]
    for linha in indicadores:
        resumo.append(linha)
    aplicar_bordas(resumo, 3, 3 + len(indicadores), 1, 2)

    por_estado: dict[str, Counter[str]] = defaultdict(Counter)
    for e in eventos:
        por_estado[e["estado_sensibilidade"]][e["tipo"]] += 1

    linha_inicio_estado = 14
    cab_estado = ["Sensibilidade ativa", "Relâmpagos", "Distúrbios", "Ruído", "Total", "% relâmpagos"]
    for col, valor in enumerate(cab_estado, start=1):
        resumo.cell(linha_inicio_estado, col, valor)
    aplicar_cabecalho(resumo, linha_inicio_estado, 1, 6)

    linhas_estado: list[list[Any]] = []
    for estado_sens, c in por_estado.items():
        total = c["Relâmpago"] + c["Distúrbio"] + c["Ruído"]
        linhas_estado.append([
            estado_sens,
            c["Relâmpago"],
            c["Distúrbio"],
            c["Ruído"],
            total,
            c["Relâmpago"] / total if total else 0,
        ])
    linhas_estado.sort(key=lambda row: (row[0] == "Desconhecida", row[0]))
    if not linhas_estado:
        linhas_estado = [["Sem deteções", 0, 0, 0, 0, 0]]

    for r, linha in enumerate(linhas_estado, start=linha_inicio_estado + 1):
        for c, valor in enumerate(linha, start=1):
            resumo.cell(r, c, valor)
        resumo.cell(r, 6).number_format = "0.00%"
    fim_estado = linha_inicio_estado + len(linhas_estado)
    aplicar_bordas(resumo, linha_inicio_estado, fim_estado, 1, 6)
    resumo.conditional_formatting.add(
        f"B{linha_inicio_estado + 1}:E{fim_estado}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="5B9BD5"),
    )

    # Dados auxiliares e gráfico.
    dados_grafico = [
        ["Tipo", "Quantidade"],
        ["Relâmpago", sum(f["relampagos"] for f in ficheiros)],
        ["Distúrbio", sum(f["disturbios"] for f in ficheiros)],
        ["Ruído", sum(f["ruido"] for f in ficheiros)],
    ]
    for r, linha in enumerate(dados_grafico, start=3):
        for c, valor in enumerate(linha, start=8):
            resumo.cell(r, c, valor)
    aplicar_cabecalho(resumo, 3, 8, 9)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Deteções por tipo"
    chart.y_axis.title = "Quantidade"
    chart.legend = None
    data = Reference(resumo, min_col=9, min_row=3, max_row=6)
    cats = Reference(resumo, min_col=8, min_row=4, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 13
    resumo.add_chart(chart, "H8")

    definir_larguras(resumo, {"A": 38, "B": 18, "C": 16, "D": 16, "E": 16, "F": 16, "H": 18, "I": 18})
    resumo.freeze_panes = "A2"

    # ---------------- Relâmpagos detalhados ----------------
    cab_eventos = [
        "Ordem", "Ficheiro", "Linha", "Data/hora do evento", "Data/hora de referência do ficheiro",
        "Segundos desde arranque", "Tipo de deteção", "Detalhe", "Distância (km)", "Sessão",
        "Noise floor", "Spike rejection", "Watchdog threshold", "Sensibilidade genérica",
        "Modo AFE", "Melhor condensador", "Ressonância (Hz)", "Estado da sensibilidade"
    ]
    relampagos_detalhados = [e for e in eventos if e["tipo"] == "Relâmpago"]
    linhas_eventos = [[
        e["ordem"], e["ficheiro"], e["linha"], e["data_hora"], e["data_referencia"],
        e["segundos"], e["tipo"], e["detalhe"], e["distancia_km"], e["sessao"],
        e["noise_floor"], e["spike_rejection"], e["watchdog_threshold"],
        e["sensibilidade_generica"], e["modo_afe"], e["melhor_condensador"],
        e["ressonancia_hz"], e["estado_sensibilidade"]
    ] for e in relampagos_detalhados]
    fim = escrever_linhas(sh_eventos, cab_eventos, linhas_eventos)
    if fim > 1:
        criar_tabela(sh_eventos, f"A1:R{fim}", "TabelaRelampagos")
        for row in range(2, fim + 1):
            sh_eventos.cell(row, 4).number_format = "yyyy-mm-dd hh:mm:ss.000"
            sh_eventos.cell(row, 5).number_format = "yyyy-mm-dd hh:mm:ss.000"
            sh_eventos.cell(row, 9).number_format = "0.000"
    sh_eventos.freeze_panes = "A2"
    definir_larguras(sh_eventos, {
        "A": 9, "B": 34, "C": 9, "D": 23, "E": 23, "F": 20, "G": 16,
        "H": 38, "I": 14, "J": 9, "K": 18, "L": 18, "M": 18, "N": 18,
        "O": 18, "P": 18, "Q": 18, "R": 42,
    })
    for row in sh_eventos.iter_rows(min_col=8, max_col=8):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    for row in sh_eventos.iter_rows(min_col=18, max_col=18):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Contagens agregadas ----------------
    cab_cont = [
        "Ficheiro", "Sessão", "Estado da sensibilidade", "Tipo de deteção", "Quantidade",
        "Primeira data/hora", "Última data/hora"
    ]
    grupos: dict[tuple[Any, ...], dict[str, Any]] = {}
    for e in eventos:
        chave = (e["ficheiro"], e["sessao"], e["estado_sensibilidade"], e["tipo"])
        reg = grupos.setdefault(chave, {"quantidade": 0, "datas": []})
        reg["quantidade"] += 1
        if e["data_hora"] is not None:
            reg["datas"].append(e["data_hora"])
    linhas_cont = []
    for chave in sorted(grupos, key=lambda x: (x[0].lower(), x[1], x[2], x[3])):
        reg = grupos[chave]
        datas = reg["datas"]
        linhas_cont.append([*chave, reg["quantidade"], min(datas) if datas else None, max(datas) if datas else None])
    fim = escrever_linhas(sh_cont, cab_cont, linhas_cont)
    if fim > 1:
        criar_tabela(sh_cont, f"A1:G{fim}", "TabelaContagens")
        for row in range(2, fim + 1):
            sh_cont.cell(row, 6).number_format = "yyyy-mm-dd hh:mm:ss.000"
            sh_cont.cell(row, 7).number_format = "yyyy-mm-dd hh:mm:ss.000"
    sh_cont.freeze_panes = "A2"
    definir_larguras(sh_cont, {"A": 36, "B": 10, "C": 42, "D": 18, "E": 14, "F": 23, "G": 23})
    for row in sh_cont.iter_rows(min_col=3, max_col=3):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Sensibilidade/configuração ----------------
    cab_sens = [
        "Ordem", "Ficheiro", "Linha", "Data/hora", "Data/hora de referência", "Sessão",
        "Categoria", "Parâmetro", "Valor anterior", "Valor novo", "Unidade", "Mensagem original",
        "Noise floor", "Spike rejection", "Watchdog threshold", "Sensibilidade genérica", "Modo AFE",
        "Melhor condensador", "Ressonância (Hz)", "Estado da sensibilidade", "TUN_CAP", "IRQ (Hz)",
        "RES teste (Hz)", "Erro (Hz)"
    ]
    linhas_sens = [[
        c["ordem"], c["ficheiro"], c["linha"], c["data_hora"], c["data_referencia"], c["sessao"],
        c["categoria"], c["parametro"], c["valor_anterior"], c["valor_novo"], c["unidade"], c["mensagem"],
        c["noise_floor"], c["spike_rejection"], c["watchdog_threshold"], c["sensibilidade_generica"],
        c["modo_afe"], c["melhor_condensador"], c["ressonancia_hz"], c["estado_sensibilidade"],
        c["tun_cap"], c["irq_hz"], c["res_teste_hz"], c["erro_hz"]
    ] for c in configs]
    fim = escrever_linhas(sh_sens, cab_sens, linhas_sens)
    if fim > 1:
        criar_tabela(sh_sens, f"A1:X{fim}", "TabelaSensibilidade")
        for row in range(2, fim + 1):
            sh_sens.cell(row, 4).number_format = "yyyy-mm-dd hh:mm:ss.000"
            sh_sens.cell(row, 5).number_format = "yyyy-mm-dd hh:mm:ss.000"
    sh_sens.freeze_panes = "A2"
    definir_larguras(sh_sens, {
        "A": 9, "B": 34, "C": 9, "D": 23, "E": 23, "F": 9, "G": 16, "H": 22,
        "I": 16, "J": 16, "K": 10, "L": 44, "M": 18, "N": 18, "O": 18, "P": 18,
        "Q": 18, "R": 18, "S": 18, "T": 42, "U": 14, "V": 14, "W": 14, "X": 14,
    })
    for col in (12, 20):
        for row in sh_sens.iter_rows(min_col=col, max_col=col):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Ficheiros ----------------
    cab_fich = [
        "Ficheiro", "Linhas", "Timestamp explícito", "Data/hora de referência", "Arranques",
        "Relâmpagos", "Distúrbios", "Ruído", "Total de deteções", "Sensibilidade no início",
        "Sensibilidade no fim", "Linhas não reconhecidas"
    ]
    linhas_fich = [[
        f["ficheiro"], f["linhas"], f["timestamp_explicito"], f["data_referencia"], f["arranques"],
        f["relampagos"], f["disturbios"], f["ruido"], f["total_deteccoes"],
        f["sensibilidade_inicio"], f["sensibilidade_fim"], f["linhas_nao_reconhecidas"]
    ] for f in ficheiros]
    fim = escrever_linhas(sh_fich, cab_fich, linhas_fich)
    if fim > 1:
        criar_tabela(sh_fich, f"A1:L{fim}", "TabelaFicheiros")
        for row in range(2, fim + 1):
            sh_fich.cell(row, 4).number_format = "yyyy-mm-dd hh:mm"
    sh_fich.freeze_panes = "A2"
    definir_larguras(sh_fich, {
        "A": 36, "B": 12, "C": 18, "D": 23, "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 14, "J": 42, "K": 42, "L": 22,
    })
    for col in (10, 11):
        for row in sh_fich.iter_rows(min_col=col, max_col=col):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Linhas não reconhecidas ----------------
    if sh_desc is not None:
        cab_desc = ["Ficheiro", "Linha", "Data/hora", "Data/hora de referência", "Sessão", "Conteúdo"]
        linhas_desc = [[
            d["ficheiro"], d["linha"], d["data_hora"], d["data_referencia"], d["sessao"], d["conteudo"]
        ] for d in desconhecidas]
        fim = escrever_linhas(sh_desc, cab_desc, linhas_desc)
        if fim > 1:
            criar_tabela(sh_desc, f"A1:F{fim}", "TabelaNaoReconhecidas")
            for row in range(2, fim + 1):
                sh_desc.cell(row, 3).number_format = "yyyy-mm-dd hh:mm:ss.000"
                sh_desc.cell(row, 4).number_format = "yyyy-mm-dd hh:mm:ss.000"
        sh_desc.freeze_panes = "A2"
        definir_larguras(sh_desc, {"A": 36, "B": 10, "C": 23, "D": 23, "E": 10, "F": 55})
        for row in sh_desc.iter_rows(min_col=6, max_col=6):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # Filtros já são incluídos pelas tabelas; ajustar visualmente todas as folhas.
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def processar_uma_pasta(pasta: Path) -> Path:
    resultado = analisar_pasta(pasta)
    destino = pasta / f"{pasta.name}.xlsx"
    return criar_excel(resultado, destino)


def pastas_com_txt(raiz: Path) -> Iterable[Path]:
    for pasta in [raiz, *sorted(p for p in raiz.rglob("*") if p.is_dir())]:
        if any(p.is_file() and p.suffix.lower() == ".txt" for p in pasta.iterdir()):
            yield pasta


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Analisa ficheiros do sensor AS3935 e cria um Excel por pasta."
    )
    parser.add_argument(
        "pasta",
        nargs="?",
        type=Path,
        default=PASTA_PREDEFINIDA,
        help=(
            "Pasta que contém os ficheiros .txt. Quando omitida, usa "
            "~/Desktop/Lightning."
        ),
    )
    parser.add_argument(
        "--cada-pasta",
        action="store_true",
        help="Processa a pasta indicada e todas as subpastas que contenham .txt",
    )
    args = parser.parse_args()

    raiz = args.pasta.expanduser().resolve()
    if not raiz.is_dir():
        print(f"Pasta inválida: {raiz}", file=sys.stderr)
        print(
            "Confirme que existe a pasta 'Lightning' no Desktop ou indique outro caminho.",
            file=sys.stderr,
        )
        return 2

    pastas = list(pastas_com_txt(raiz)) if args.cada_pasta else [raiz]
    if not pastas:
        print("Não foram encontradas pastas com ficheiros .txt.", file=sys.stderr)
        return 1

    houve_erro = False
    for pasta in pastas:
        try:
            destino = processar_uma_pasta(pasta)
            print(f"Criado: {destino}")
        except Exception as exc:
            houve_erro = True
            print(f"Erro em {pasta}: {exc}", file=sys.stderr)

    return 1 if houve_erro else 0


if __name__ == "__main__":
    raise SystemExit(main())